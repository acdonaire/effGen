"""
Excel parsing tool for the effGen framework.

Primary backend : openpyxl (reads .xlsx).
DataFrame output: pandas (optional, used when as_dataframe=True).

Operations
----------
- sheets      : List all sheet names in the workbook.
- read_sheet  : Read a sheet as a list-of-rows or a pandas DataFrame.
- headers     : Return the header row of a given sheet.
- metadata    : Return workbook properties (author, title, created, etc.).

Input: file path (str/Path) OR raw XLSX bytes.
.xls files are not supported by openpyxl; callers should convert to .xlsx first.
Output: structured dicts with success/data/error keys.
"""

from __future__ import annotations

import asyncio
import io
import logging
from pathlib import Path
from typing import Any

from effgen.errors import CorruptDocumentError

from ..base_tool import (
    BaseTool,
    ParameterSpec,
    ParameterType,
    ToolCategory,
    ToolMetadata,
)
from ._fs import confine_path, normalize_allowed_dirs

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Source resolution
# ---------------------------------------------------------------------------

def _resolve_xlsx(source: str | bytes | Path) -> bytes:
    if isinstance(source, str | Path):
        p = Path(source)
        if not p.exists():
            raise FileNotFoundError(f"Excel file not found: {source}")
        lower = str(source).lower()
        if lower.endswith(".xls") and not lower.endswith(".xlsx"):
            raise ValueError(
                "Legacy .xls format is not supported. "
                "Please convert to .xlsx (use LibreOffice, Excel, or xlrd+openpyxl)."
            )
        return p.read_bytes()
    if isinstance(source, bytes | bytearray):
        return bytes(source)
    raise TypeError(f"Expected str, Path, or bytes; got {type(source)}")


def _open_workbook(raw: bytes):
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for ExcelTool. "
            "Install with: pip install 'effgen[documents]'"
        ) from exc
    try:
        return openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as exc:
        raise CorruptDocumentError("xlsx", str(exc)) from exc


# ---------------------------------------------------------------------------
# Operation implementations
# ---------------------------------------------------------------------------

def _excel_sheets(source: str | bytes | Path) -> dict[str, Any]:
    raw = _resolve_xlsx(source)
    wb = _open_workbook(raw)
    names = wb.sheetnames
    return {
        "success": True,
        "data": {"sheets": names, "count": len(names)},
        "sheets": names,
        "count": len(names),
        "error": None,
    }


def _excel_read_sheet(
    source: str | bytes | Path,
    sheet_name: str | None = None,
    as_dataframe: bool = False,
) -> dict[str, Any]:
    raw = _resolve_xlsx(source)
    wb = _open_workbook(raw)

    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))

    if as_dataframe:
        try:
            import pandas as pd
            if rows:
                headers = rows[0]
                data_rows = rows[1:]
                df = pd.DataFrame(data_rows, columns=headers)
                records = df.to_dict(orient="records")
            else:
                records = []
            return {
                "success": True,
                "data": {
                    "sheet": sheet_name,
                    "records": records,
                    "row_count": len(records),
                    "columns": list(df.columns) if rows else [],
                },
                "sheet": sheet_name,
                "records": records,
                "row_count": len(records),
                "columns": list(df.columns) if rows else [],
                "error": None,
            }
        except ImportError:
            logger.warning("pandas not installed; falling back to raw rows output.")

    return {
        "success": True,
        "data": {
            "sheet": sheet_name,
            "rows": rows,
            "row_count": len(rows),
            "col_count": max((len(r) for r in rows), default=0),
        },
        "sheet": sheet_name,
        "rows": rows,
        "row_count": len(rows),
        "col_count": max((len(r) for r in rows), default=0),
        "error": None,
    }


def _excel_headers(source: str | bytes | Path, sheet_name: str | None = None) -> dict[str, Any]:
    raw = _resolve_xlsx(source)
    wb = _open_workbook(raw)

    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [str(h) if h is not None else "" for h in first_row]

    return {
        "success": True,
        "data": {"sheet": sheet_name, "headers": headers, "count": len(headers)},
        "sheet": sheet_name,
        "headers": headers,
        "count": len(headers),
        "error": None,
    }


def _excel_metadata(source: str | bytes | Path) -> dict[str, Any]:
    raw = _resolve_xlsx(source)
    wb = _open_workbook(raw)
    props = wb.properties
    info: dict[str, Any] = {
        "title": props.title,
        "subject": props.subject,
        "author": props.creator,
        "keywords": props.keywords,
        "description": props.description,
        "last_modified_by": props.lastModifiedBy,
        "created": str(props.created) if props.created else None,
        "modified": str(props.modified) if props.modified else None,
        "category": props.category,
        "sheet_names": wb.sheetnames,
        "sheet_count": len(wb.sheetnames),
    }
    return {
        "success": True,
        "data": info,
        "metadata": info,
        "error": None,
    }


# ---------------------------------------------------------------------------
# BaseTool subclass
# ---------------------------------------------------------------------------

class ExcelTool(BaseTool):
    """
    Parse Excel XLSX workbooks.

    Primary backend: openpyxl. Pandas used for DataFrame output when available.

    Accepts file paths (str/Path) OR raw XLSX bytes.
    Raises CorruptDocumentError on malformed input.
    Note: .xls (legacy Excel) is not supported; convert to .xlsx first.
    """

    def __init__(self, allowed_directories: list[str] | None = None) -> None:
        """Args:
            allowed_directories: Roots a file path may be read from. By default
                any path is allowed except protected system and credential
                locations (/etc, /proc, ~/.ssh, cloud creds, …), which are
                always refused. Pass a list to confine reads to those roots only.
        """
        self._allowed_dirs = normalize_allowed_dirs(allowed_directories)
        super().__init__(
            metadata=ToolMetadata(
                name="excel",
                description=(
                    "Parse Excel XLSX workbooks. Operations: sheets (list sheet names), "
                    "read_sheet (read data as rows or DataFrame), "
                    "headers (get column headers), metadata (workbook properties). "
                    "Accepts file paths or raw bytes. .xls not supported; use .xlsx."
                ),
                category=ToolCategory.DATA_PROCESSING,
                parameters=[
                    ParameterSpec(
                        name="operation",
                        type=ParameterType.STRING,
                        description="Operation: 'sheets', 'read_sheet', 'headers', or 'metadata'.",
                        required=True,
                        enum=["sheets", "read_sheet", "headers", "metadata"],
                    ),
                    ParameterSpec(
                        name="path",
                        type=ParameterType.STRING,
                        description="Path to the XLSX file.",
                        required=False,
                    ),
                    ParameterSpec(
                        name="sheet_name",
                        type=ParameterType.STRING,
                        description="Sheet name. Defaults to the first sheet.",
                        required=False,
                    ),
                    ParameterSpec(
                        name="as_dataframe",
                        type=ParameterType.BOOLEAN,
                        description=(
                            "For read_sheet: if True, return data as pandas-style records "
                            "(list of dicts keyed by header). Requires pandas."
                        ),
                        required=False,
                        default=False,
                    ),
                ],
                timeout_seconds=30,
                tags=["excel", "xlsx", "spreadsheet", "parsing", "tables"],
                examples=[
                    {"operation": "sheets", "path": "/path/to/data.xlsx"},
                    {
                        "operation": "read_sheet",
                        "path": "/path/to/data.xlsx",
                        "sheet_name": "Sheet1",
                        "as_dataframe": True,
                    },
                    {"operation": "headers", "path": "/path/to/data.xlsx"},
                    {"operation": "metadata", "path": "/path/to/data.xlsx"},
                ],
            )
        )

    async def _execute(
        self,
        operation: str,
        path: str | None = None,
        xlsx_bytes: bytes | None = None,
        sheet_name: str | None = None,
        as_dataframe: bool = False,
        **kwargs: Any,
    ) -> dict[str, Any]:
        source: str | bytes | Path
        if path:
            source = str(confine_path(path, self._allowed_dirs))
        elif xlsx_bytes is not None:
            source = xlsx_bytes
        else:
            raise ValueError("Provide 'path' or 'xlsx_bytes'.")

        op = operation.lower()

        if op == "sheets":
            return await asyncio.to_thread(_excel_sheets, source)
        elif op == "read_sheet":
            return await asyncio.to_thread(_excel_read_sheet, source, sheet_name, as_dataframe)
        elif op == "headers":
            return await asyncio.to_thread(_excel_headers, source, sheet_name)
        elif op == "metadata":
            return await asyncio.to_thread(_excel_metadata, source)
        else:
            raise ValueError(
                f"Unknown operation: {operation!r}. "
                "Use 'sheets', 'read_sheet', 'headers', or 'metadata'."
            )
